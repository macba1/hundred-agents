#!/usr/bin/env python3
"""Importa la agenda de clientes como versión inmutable.

Es la fuente oficial de **qué negocios pueden identificarse solos** por
WhatsApp, igual que `Tarifas todas.xlsx` lo es de los precios. Aquí no se
inventa ningún cliente: si un negocio no está en este fichero, no es cliente
de Chacón y el sistema no puede convertirlo en uno.

Lo que se comprobó del fichero antes de modelarlo:

  - 230 filas · 207 códigos · 207 razones sociales · 19 filas con centro
  - **código y razón social son 1:1**: ningún código tiene dos nombres y
    ningún nombre tiene dos códigos. Por eso el cliente se identifica por
    código y el centro es un atributo suyo, no otro cliente.
  - 9 filas exactamente repetidas, que no deben salir dos veces al buscar
  - 11 códigos con más de un valor de centro; en casi todos es "un centro y
    otra fila sin centro", como `50146 BOLLYSUR` (`01` y vacío)
  - los centros mezclan formatos: conviven `01` y `1`, `03` y `3`, en
    clientes distintos. **No son equivalentes** y se guardan tal cual

El código y el centro son IDENTIFICADORES, no cantidades: se guardan como
texto y `01` nunca se convierte en `1`. Y el vínculo teléfono→cliente se hace
contra código+centro, nunca contra el número de fila, para que una agenda
nueva no rompa los vínculos ya establecidos.

    python3 chacon-alcantara/import/extraer_clientes.py \
        --xlsx "~/Downloads/Agenda clientes por ruta.xlsx"
    python3 chacon-alcantara/import/extraer_clientes.py --aprobar 1 --por "Nombre"
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[2]
DATOS = RAIZ / "chacon-alcantara" / "data"
AGENDA = DATOS / "clientes"

# Sufijos societarios: se quitan para BUSCAR, nunca del dato guardado.
SUFIJOS = [
    "s l u", "s l l", "s coop and", "s coop", "s c a", "s a u",
    "sociedad limitada", "sociedad anonima",
    "s l", "s a", "c b", "s c", "scp", "sl", "sa", "cb",
]


def sin_tildes(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def normalizar(nombre: str) -> str:
    """Forma canónica para comparar. El original NO se toca."""
    s = sin_tildes(nombre).lower()
    s = re.sub(r"[^a-z0-9ñ ]+", " ", s)          # comas, puntos, comillas
    s = re.sub(r"\s+", " ", s).strip()
    return s


def sin_sufijo(normalizado: str) -> str:
    """Quita el sufijo societario del final, si lo hay."""
    s = normalizado
    cambiado = True
    while cambiado:
        cambiado = False
        for suf in SUFIJOS:
            if s.endswith(f" {suf}"):
                s = s[: -(len(suf) + 1)].strip()
                cambiado = True
                break
    return s


def leer(xlsx: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("Falta openpyxl. Instala con: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = []
    for n, valores in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        codigo, centro, razon = (tuple(valores) + (None,) * 3)[:3]
        if codigo is None and razon is None:
            continue
        filas.append({
            # Identificadores: SIEMPRE texto. `01` no puede volverse `1`.
            "customer_code": str(codigo).strip() if codigo is not None else None,
            "customer_center": (str(centro).strip()
                                if centro is not None and str(centro).strip() else None),
            "legal_name": str(razon).strip() if razon is not None else None,
            "source_row": n,
        })
    return filas


def construir(filas: list[dict]) -> tuple[list[dict], dict]:
    """Agrupa filas en clientes. Un cliente = un código."""
    por_codigo: dict[str, dict] = {}
    vistas = set()
    duplicados_exactos = 0
    conflictos_nombre = []

    for f in filas:
        clave = (f["customer_code"], f["customer_center"], f["legal_name"])
        if clave in vistas:
            duplicados_exactos += 1
            continue
        vistas.add(clave)

        cod = f["customer_code"]
        if cod not in por_codigo:
            por_codigo[cod] = {
                "customer_code": cod,
                "legal_name": f["legal_name"],
                # Sin nombre comercial todavía: se rellena desde el panel.
                "display_name": f["legal_name"],
                "aliases": [],
                "centers": [],
                "search_normalized": normalizar(f["legal_name"]),
                "search_sin_sufijo": sin_sufijo(normalizar(f["legal_name"])),
                "source_rows": [],
                "status": "activo",
            }
        c = por_codigo[cod]
        if c["legal_name"] != f["legal_name"]:
            conflictos_nombre.append((cod, c["legal_name"], f["legal_name"]))
        if f["customer_center"] not in c["centers"]:
            c["centers"].append(f["customer_center"])
        c["source_rows"].append(f["source_row"])

    clientes = sorted(por_codigo.values(), key=lambda x: x["customer_code"])
    for c in clientes:
        # `None` = sin centro informado. Se ordena al final, y NUNCA se
        # convierte en `0` ni en `"0"`: son cosas distintas.
        c["centers"].sort(key=lambda x: (x is None, x or ""))
        c["multi_centro"] = len([x for x in c["centers"] if x]) > 1

    resumen = {
        "filas_fuente": len(filas),
        "duplicados_exactos": duplicados_exactos,
        "filas_unicas": len(vistas),
        "clientes_unicos": len(clientes),
        "con_centro": sum(1 for c in clientes if any(c["centers"])),
        "multi_centro": sum(1 for c in clientes if c["multi_centro"]),
        "conflictos_nombre": conflictos_nombre,
        "centros_vistos": sorted({x for c in clientes for x in c["centers"] if x}),
    }
    return clientes, resumen


def invariantes(clientes: list[dict], resumen: dict) -> list[str]:
    fallos = []
    if resumen["conflictos_nombre"]:
        fallos.append(f"{len(resumen['conflictos_nombre'])} códigos con más de una razón social")

    por_nombre = defaultdict(set)
    for c in clientes:
        por_nombre[c["legal_name"]].add(c["customer_code"])
    repes = {k: v for k, v in por_nombre.items() if len(v) > 1}
    if repes:
        fallos.append(f"{len(repes)} razones sociales con más de un código: "
                      f"{list(repes.items())[:3]}")

    for c in clientes:
        if not c["customer_code"] or not str(c["customer_code"]).strip():
            fallos.append("hay un cliente sin código")
        if not isinstance(c["customer_code"], str):
            fallos.append(f"código que no es texto: {c['customer_code']!r}")
        for cen in c["centers"]:
            if cen is not None and not isinstance(cen, str):
                fallos.append(f"centro que no es texto en {c['customer_code']}: {cen!r}")
        if not c["legal_name"]:
            fallos.append(f"cliente {c['customer_code']} sin razón social")

    # Los ceros iniciales tienen que haber sobrevivido.
    con_cero = [x for x in resumen["centros_vistos"] if x.startswith("0")]
    if resumen["centros_vistos"] and not con_cero:
        fallos.append("ningún centro conserva el cero inicial: revisa la lectura del Excel")
    return fallos


def estado() -> dict:
    ruta = AGENDA / "estado.json"
    if ruta.exists():
        return json.loads(ruta.read_text(encoding="utf-8"))
    return {"version_activa": None, "versiones": []}


def guardar_estado(e: dict) -> None:
    AGENDA.mkdir(parents=True, exist_ok=True)
    (AGENDA / "estado.json").write_text(json.dumps(e, ensure_ascii=False, indent=2),
                                        encoding="utf-8")


def cargar_version(n: int) -> dict | None:
    ruta = AGENDA / f"version-{n}.json"
    return json.loads(ruta.read_text(encoding="utf-8")) if ruta.exists() else None


def diff(nuevos: list[dict], anteriores: list[dict] | None) -> dict:
    """Diff por CÓDIGO, nunca por número de fila."""
    if anteriores is None:
        return {"sin_version_anterior": True, "altas": len(nuevos), "bajas": [],
                "cambios_de_nombre": [], "cambios_de_centros": []}
    a = {c["customer_code"]: c for c in anteriores}
    b = {c["customer_code"]: c for c in nuevos}
    bajas = sorted(set(a) - set(b))
    return {
        "sin_version_anterior": False,
        "altas": len(set(b) - set(a)),
        "detalle_altas": sorted(set(b) - set(a))[:20],
        # Un cliente que desaparece NO se borra: se marca para que lo revisen.
        "bajas": [{"customer_code": k, "legal_name": a[k]["legal_name"],
                   "estado": "SOURCE_MISSING"} for k in bajas],
        "cambios_de_nombre": [
            {"customer_code": k, "antes": a[k]["legal_name"], "ahora": b[k]["legal_name"]}
            for k in sorted(set(a) & set(b)) if a[k]["legal_name"] != b[k]["legal_name"]],
        "cambios_de_centros": [
            {"customer_code": k, "antes": a[k]["centers"], "ahora": b[k]["centers"]}
            for k in sorted(set(a) & set(b)) if a[k]["centers"] != b[k]["centers"]],
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx")
    ap.add_argument("--aprobar", type=int, metavar="N")
    ap.add_argument("--por")
    ap.add_argument("--listar", action="store_true")
    a = ap.parse_args()

    AGENDA.mkdir(parents=True, exist_ok=True)
    e = estado()

    if a.listar:
        print(f"Versión activa: {e['version_activa'] or '(ninguna)'}")
        for v in e["versiones"]:
            marca = "ACTIVA" if v["version"] == e["version_activa"] else \
                    ("aprobada" if v.get("approved") else "pendiente")
            print(f"  v{v['version']}  {v['generado'][:19]}  {v['clientes_unicos']} clientes  [{marca}]")
        return

    if a.aprobar is not None:
        if not a.por:
            sys.exit("Aprobar exige --por: queda registrado quién lo hizo.")
        v = cargar_version(a.aprobar)
        if not v:
            sys.exit(f"No existe la versión {a.aprobar}.")
        if v["invariantes_fallidos"]:
            sys.exit(f"La versión {a.aprobar} no pasa sus invariantes.")
        v["approved"] = True
        v["approved_by"] = a.por
        v["approved_at"] = datetime.now(timezone.utc).isoformat()
        (AGENDA / f"version-{a.aprobar}.json").write_text(
            json.dumps(v, ensure_ascii=False, indent=2), encoding="utf-8")
        e["version_activa"] = a.aprobar
        for x in e["versiones"]:
            if x["version"] == a.aprobar:
                x["approved"] = True
                x["approved_by"] = a.por
        e.setdefault("auditoria", []).append({
            "ts": datetime.now(timezone.utc).isoformat(),
            "accion": "activar_agenda", "version": a.aprobar, "por": a.por})
        guardar_estado(e)
        print(f"Agenda v{a.aprobar} aprobada y activada por {a.por}.")
        if v["diff_contra_activa"].get("bajas"):
            print(f"OJO: {len(v['diff_contra_activa']['bajas'])} clientes ya no están en la "
                  "agenda nueva. Quedan marcados SOURCE_MISSING para revisión, no se borran.")
        return

    if not a.xlsx:
        sys.exit("Hace falta --xlsx, --aprobar o --listar.")

    xlsx = Path(a.xlsx).expanduser()
    sha = hashlib.sha256(xlsx.read_bytes()).hexdigest()
    version = max([v["version"] for v in e["versiones"]], default=0) + 1

    filas = leer(xlsx)
    clientes, resumen = construir(filas)
    fallos = invariantes(clientes, resumen)
    activa = cargar_version(e["version_activa"]) if e["version_activa"] else None
    d = diff(clientes, activa["clientes"] if activa else None)

    doc = {
        "version": version,
        "generado": datetime.now(timezone.utc).isoformat(),
        "source_file": xlsx.name,
        "source_sha256": sha,
        "approved": False,
        "invariantes_fallidos": fallos,
        "resumen": resumen,
        "diff_contra_activa": d,
        "clientes": clientes,
        # La fuente cruda se conserva para poder auditar cualquier duda.
        "filas_fuente": filas,
    }
    (AGENDA / f"version-{version}.json").write_text(
        json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    e["versiones"].append({
        "version": version, "generado": doc["generado"], "source_file": xlsx.name,
        "source_sha256": sha, "clientes_unicos": resumen["clientes_unicos"],
        "filas_fuente": resumen["filas_fuente"], "approved": False,
        "invariantes_fallidos": len(fallos)})
    guardar_estado(e)

    print(f"Agenda v{version} desde {xlsx.name}")
    print(f"  filas en el fichero      {resumen['filas_fuente']}")
    print(f"  duplicados exactos       {resumen['duplicados_exactos']}")
    print(f"  filas únicas             {resumen['filas_unicas']}")
    print(f"  CLIENTES únicos          {resumen['clientes_unicos']}")
    print(f"  con centro informado     {resumen['con_centro']}")
    print(f"  con VARIOS centros       {resumen['multi_centro']}")
    print(f"  centros vistos           {resumen['centros_vistos']}")
    print()
    print("Invariantes: " + ("los 5 en verde." if not fallos else f"{len(fallos)} FALLIDOS"))
    for f in fallos[:10]:
        print(f"  ✗ {f}")
    print()
    if d["sin_version_anterior"]:
        print("Diff: primera versión, todo son altas.")
    else:
        print(f"Diff: {d['altas']} altas · {len(d['bajas'])} desaparecidos · "
              f"{len(d['cambios_de_nombre'])} cambios de nombre · "
              f"{len(d['cambios_de_centros'])} cambios de centros")
    print()
    print("PENDIENTE. No cambia nada hasta:")
    print(f"  python3 chacon-alcantara/import/extraer_clientes.py --aprobar {version} --por 'Nombre'")


if __name__ == "__main__":
    main()
