#!/usr/bin/env python3
"""Importa las tarifas de Chacón como una versión INMUTABLE.

Acepta el Excel (`--xlsx`) o el PDF (`--pdf`). **El Excel es la fuente oficial
de precios**; del PDF salen descripción, marca, unidades por caja, peso, EAN,
alérgenos, observaciones e imágenes, y su columna de precio NO se usa para
calcular pedidos. Los dos producen exactamente el mismo formato de versión,
así que el resto del sistema no se entera de cuál se usó.

Por qué versionado y no sobrescrito: un pedido confirmado guarda el precio que
se le dijo al cliente. Si una reimportación pisara la tabla activa, los pedidos
de ayer cambiarían de importe solos. Aquí cada importación crea una versión
nueva, enseña el diff y **espera aprobación**. La activa no se toca.

El dinero NO se guarda en float. Los precios llegan a cuatro decimales
(`0,0001`) y un float los arruina en la primera suma. Se guardan como enteros
en diezmilésimas de euro: `13,889 €` -> `138890`. Se redondea a céntimos solo
al calcular el importe final de una línea.

Invariantes que se comprueban en cada importación, y que la abortan si fallan:
  649 registros · 133 códigos por tarifa normal · conjuntos 1=2=3=4 ·
  19 por tarifa de oferta · conjuntos 1OF=2OF=3OF=4OF · sin duplicados de
  (tariff_code, product_code) · ceros iniciales intactos · coma decimal
  convertida · ninguna oferta >= su precio normal del mismo nivel.

    python3 chacon-alcantara/import/extraer_tarifas.py \
        --pdf "~/Downloads/Tarifas todas.pdf"
    python3 chacon-alcantara/import/extraer_tarifas.py --aprobar 1
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

try:
    import pymupdf
except ImportError:  # pragma: no cover
    try:
        import fitz as pymupdf
    except ImportError:
        sys.exit("Falta PyMuPDF. Instala con: pip install pymupdf")

RAIZ = Path(__file__).resolve().parents[2]
DATOS = RAIZ / "chacon-alcantara" / "data"
TARIFAS = DATOS / "tarifas"

ESCALA = 10_000            # diezmilésimas de euro

CABECERAS = {"Tarifas", "Código", "Nombre", "Precio", "Tarifa", "NOMBRE TARIFA"}
RE_FECHA = re.compile(r"^\w+,\s+\d+\s+de\s+\w+\s+de\s+\d{4}$")
RE_PRECIO = re.compile(r"^-?\d{1,8}(?:,\d{1,4})?$")
RE_SECCION = re.compile(r"^(?:[1-4](?:OF)?|ALI|COO|OFC|S)$")

# tariff_code -> (tier, price_type, etiqueta esperada)
TARIFAS_CONOCIDAS = {
    "1":   ("1", "standard", "PIEZA"),
    "1OF": ("1", "offer",    "PIEZA OFERTA"),
    "2":   ("2", "standard", "1/2 CAJA"),
    "2OF": ("2", "offer",    "1/2 CAJA OFERTA"),
    "3":   ("3", "standard", "1 CAJA"),
    "3OF": ("3", "offer",    "1 CAJA OFERTA"),
    "4":   ("4", "standard", "+ 2 CAJAS"),
    "4OF": ("4", "offer",    "+ 2 CAJAS OFERTA"),
    "ALI": ("special", "special", "ALIPENSA"),
    "COO": ("special", "special", "COOPERATIVA VINO"),
    # El PDF recorta esta etiqueta a 17 caracteres; el Excel la trae entera.
    "OFC": ("special", "special", "OFERTA COOPERATIVA"),
    "S":   ("special", "special", "SPV"),
}

NORMALES = ["1", "2", "3", "4"]
OFERTAS = ["1OF", "2OF", "3OF", "4OF"]
ESPECIALES = ["ALI", "COO", "OFC", "S"]

# Artículos que existen en la tarifa pero no son producto para una tienda.
# No se borran —el importador no decide qué vende Chacón— pero salen marcados.
RE_INTERNOS = re.compile(
    r"^(PORTADA|10000|9995|4525|87E\d+|B\d{5,}|FCOM\d+|6010000016|901[0-9])$")
PALABRAS_INTERNAS = re.compile(
    r"portada|articulo por si no sale|portes|palet|etiqueta|bateria|"
    r"rollo film|rollo plastico|bolsas\s+vacio", re.I)


def a_entero(precio: str) -> int:
    """`13,889` o `13.889` -> 138890. Exacto, sin float por medio.

    El PDF usa coma decimal y el Excel punto. Se acepta cualquiera de los dos,
    pero NUNCA los dos a la vez en el mismo número: `1.234,56` sería un
    separador de miles y aquí no aparece, así que se rechaza en vez de
    adivinar."""
    p = str(precio).strip()
    if "," in p and "." in p:
        raise ValueError(f"precio con coma Y punto, ambiguo: {p}")
    d = Decimal(p.replace(",", "."))
    escalado = (d * ESCALA).normalize()
    if escalado != escalado.to_integral_value():
        raise ValueError(f"precio con más de 4 decimales: {precio}")
    return int(escalado)


def de_entero(n: int) -> str:
    """138890 -> '13,889' (sin ceros de relleno). Solo para mostrar."""
    d = (Decimal(n) / ESCALA).normalize()
    return format(d, "f").replace(".", ",")


def parsear_xlsx(xlsx_path: Path) -> tuple[list[dict], dict]:
    """Lee el Excel. Viene agrupado: una fila con el código de tarifa en la
    primera columna y debajo sus artículos.

    Los códigos llegan con espacios de relleno y los precios con punto
    decimal, al revés que en el PDF. Se recortan y se normalizan aquí, sin
    tocar los ceros iniciales: `0052` no puede volverse `52`."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Falta openpyxl. Instala con: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas: list[dict] = []
    tarifa = None
    for fila, valores in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        grupo, codigo, nombre, precio, etiqueta = (tuple(valores) + (None,) * 5)[:5]
        if grupo is not None and str(grupo).strip():
            posible = str(grupo).strip()
            if RE_SECCION.match(posible):
                tarifa = posible
            continue
        if codigo is None or precio is None:
            continue
        filas.append({
            "product_code": str(codigo).strip(),      # string, ceros incluidos
            "product_name": re.sub(r"\s+", " ", str(nombre or "")).strip(),
            "precio_bruto": str(precio).strip(),
            "tier_label": re.sub(r"\s+", " ", str(etiqueta or "")).strip() or None,
            "tariff_code": tarifa,
            "source_page": fila,                       # fila del Excel
        })
    return filas, {"hoja": ws.title, "filas_totales": ws.max_row,
                   "origen": "xlsx"}


def parsear(pdf_path: Path) -> tuple[list[dict], dict]:
    doc = pymupdf.open(pdf_path)
    filas: list[dict] = []
    tarifa = None
    paginas_con_contenido = 0

    for pagina in range(1, doc.page_count + 1):
        texto = doc[pagina - 1].get_text()
        if not texto.strip():
            continue
        paginas_con_contenido += 1
        lineas = [l.strip() for l in texto.split("\n")]
        lineas = [l for l in lineas
                  if l and l not in CABECERAS and not RE_FECHA.match(l)]

        i = 0
        while i < len(lineas):
            if RE_SECCION.match(lineas[i]):
                tarifa = lineas[i]
                i += 1
                continue
            # Una fila es: código / nombre / precio / etiqueta de tramo.
            if i + 2 < len(lineas) and RE_PRECIO.match(lineas[i + 2]):
                siguiente = lineas[i + 3] if i + 3 < len(lineas) else None
                etiqueta = (siguiente if siguiente and not RE_PRECIO.match(siguiente)
                            and not RE_SECCION.match(siguiente) else None)
                filas.append({
                    "product_code": lineas[i],          # SIEMPRE string
                    "product_name": lineas[i + 1],
                    "precio_bruto": lineas[i + 2],
                    "tier_label": etiqueta,
                    "tariff_code": tarifa,
                    "source_page": pagina,
                })
                i += 4 if etiqueta else 3
                continue
            i += 1

    return filas, {"paginas_fisicas": doc.page_count,
                   "paginas_con_contenido": paginas_con_contenido,
                   "paginas_en_blanco": doc.page_count - paginas_con_contenido}


def comprobar_invariantes(filas: list[dict]) -> list[str]:
    """Devuelve la lista de fallos. Vacía = la importación es fiable."""
    fallos = []
    por = defaultdict(dict)
    for f in filas:
        por[f["tariff_code"]][f["product_code"]] = f

    if len(filas) != 649:
        fallos.append(f"registros: {len(filas)}, esperados 649")

    for t in NORMALES:
        if len(por[t]) != 133:
            fallos.append(f"tarifa {t}: {len(por[t])} códigos, esperados 133")
    base = set(por["1"])
    for t in ("2", "3", "4"):
        if set(por[t]) != base:
            fallos.append(f"la tarifa {t} no tiene los mismos códigos que la 1: "
                          f"{sorted(set(por[t]) ^ base)[:5]}")

    for t in OFERTAS:
        if len(por[t]) != 19:
            fallos.append(f"tarifa {t}: {len(por[t])} códigos, esperados 19")
    baseof = set(por["1OF"])
    for t in ("2OF", "3OF", "4OF"):
        if set(por[t]) != baseof:
            fallos.append(f"la tarifa {t} no coincide con 1OF: {sorted(set(por[t]) ^ baseof)[:5]}")
    if not baseof <= base:
        fallos.append(f"hay ofertas sin tarifa normal: {sorted(baseof - base)}")

    vistos = set()
    for f in filas:
        clave = (f["tariff_code"], f["product_code"])
        if clave in vistos:
            fallos.append(f"duplicado {clave}")
        vistos.add(clave)

    for t in filas:
        if t["tariff_code"] not in TARIFAS_CONOCIDAS:
            fallos.append(f"tarifa desconocida: {t['tariff_code']}")
        else:
            esperada = TARIFAS_CONOCIDAS[t["tariff_code"]][2]
            visto = re.sub(r"\s+", " ", str(t["tier_label"] or "")).strip()
            # El PDF recorta las etiquetas al ancho de su columna, así que un
            # prefijo de la esperada es un truncamiento, no un error. Lo que no
            # se admite es una etiqueta distinta: eso sería otra tarifa.
            if visto != esperada and not esperada.startswith(visto):
                fallos.append(f"{t['product_code']} en {t['tariff_code']}: etiqueta "
                              f"'{t['tier_label']}' en vez de '{esperada}'")

    # Ceros iniciales: si alguno se hubiera vuelto número, se habría perdido.
    for f in filas:
        if not isinstance(f["product_code"], str):
            fallos.append(f"código que no es string: {f['product_code']!r}")

    # Ninguna oferta puede igualar ni superar su normal del mismo nivel.
    for normal, oferta in zip(NORMALES, OFERTAS):
        for cod, fo in por[oferta].items():
            fn = por[normal].get(cod)
            if not fn:
                fallos.append(f"{cod} tiene {oferta} pero no {normal}")
                continue
            if a_entero(fo["precio_bruto"]) >= a_entero(fn["precio_bruto"]):
                fallos.append(f"{cod}: oferta {fo['precio_bruto']} >= normal "
                              f"{fn['precio_bruto']} en el nivel {normal}")
    return fallos


def normalizar(filas: list[dict], pdf_path: Path, version: int, sha: str) -> list[dict]:
    out = []
    for f in filas:
        tier, tipo, _ = TARIFAS_CONOCIDAS[f["tariff_code"]]
        interno = bool(RE_INTERNOS.match(f["product_code"])
                       or PALABRAS_INTERNAS.search(f["product_name"]))
        # El prefijo OF de un CÓDIGO no es una oferta: son artículos distintos.
        # Las ofertas son las TABLAS 1OF/2OF/3OF/4OF. Confundirlo regalaría
        # producto o cobraría 0,001 € como precio comercial.
        codigo_of = f["product_code"].upper().startswith("OF")
        out.append({
            "product_code": f["product_code"],
            "product_name": re.sub(r"\s+", " ", f["product_name"]).strip(),
            "tariff_code": f["tariff_code"],
            "tier": tier,
            "price_type": tipo,
            "price_e4": a_entero(f["precio_bruto"]),      # entero, escala 1e-4
            "price_display": de_entero(a_entero(f["precio_bruto"])),
            "price_raw": f["precio_bruto"],
            "tier_label": f["tier_label"],
            "source_file": pdf_path.name,
            "source_page": f["source_page"],
            "catalog_version": version,
            "source_sha256": sha,
            "approved": False,
            "active": False,
            "valid_from": None,
            "valid_until": None,
            "es_articulo_interno": interno,
            "codigo_empieza_por_of": codigo_of,
            "review_status": ("promotion_rule_required" if codigo_of
                              else "internal_review_required" if interno else None),
        })
    return out


def estado() -> dict:
    ruta = TARIFAS / "estado.json"
    if ruta.exists():
        return json.loads(ruta.read_text(encoding="utf-8"))
    return {"version_activa": None, "versiones": []}


def guardar_estado(e: dict) -> None:
    TARIFAS.mkdir(parents=True, exist_ok=True)
    (TARIFAS / "estado.json").write_text(json.dumps(e, ensure_ascii=False, indent=2),
                                         encoding="utf-8")


def cargar_version(n: int) -> dict | None:
    ruta = TARIFAS / f"version-{n}.json"
    return json.loads(ruta.read_text(encoding="utf-8")) if ruta.exists() else None


def diff(nueva: list[dict], anterior: list[dict] | None) -> dict:
    """Diff por (tariff_code, product_code). Sin anterior, todo es nuevo."""
    if anterior is None:
        return {"sin_version_anterior": True, "altas": len(nueva),
                "bajas": 0, "cambios_de_precio": [], "detalle_altas": []}
    idx = lambda fs: {(f["tariff_code"], f["product_code"]): f for f in fs}
    a, b = idx(anterior), idx(nueva)
    altas = sorted(set(b) - set(a))
    bajas = sorted(set(a) - set(b))
    cambios = []
    for k in sorted(set(a) & set(b)):
        if a[k]["price_e4"] != b[k]["price_e4"]:
            cambios.append({"tariff_code": k[0], "product_code": k[1],
                            "antes": a[k]["price_display"], "ahora": b[k]["price_display"],
                            "direccion": "sube" if b[k]["price_e4"] > a[k]["price_e4"] else "baja"})
    return {"sin_version_anterior": False, "altas": len(altas), "bajas": len(bajas),
            "cambios_de_precio": cambios,
            "detalle_altas": [{"tariff_code": t, "product_code": c} for t, c in altas],
            "detalle_bajas": [{"tariff_code": t, "product_code": c} for t, c in bajas]}


def legado(filas: list[dict]) -> dict:
    """Cruza el catálogo anterior con la nueva tarifa.

    Los precios antiguos que no aparecen en NINGUNA tabla del PDF nuevo no se
    borran: se guardan como evidencia histórica, inactivos y marcados para
    revisión. Borrarlos sería perder la única prueba de qué se cobró antes.
    """
    ruta = DATOS / "catalogo-normalizado.json"
    if not ruta.exists():
        return {"sin_catalogo_anterior": True}
    viejo = json.loads(ruta.read_text(encoding="utf-8"))

    todos_precios = defaultdict(set)
    for f in filas:
        todos_precios[f["product_code"]].add(f["price_e4"])

    por_codigo = defaultdict(list)
    for p in viejo["productos"]:
        if p.get("tarifa") is not None:
            por_codigo[p["codigo"]].append(p["_original"]["tarifa"])

    huerfanos, resueltos = [], 0
    for cod, precios in sorted(por_codigo.items()):
        for bruto in precios:
            e4 = a_entero(bruto)
            if e4 in todos_precios.get(cod, set()):
                resueltos += 1
            else:
                huerfanos.append({
                    "product_code": cod, "price_raw": bruto, "price_e4": e4,
                    "price_display": de_entero(e4),
                    "estado": "legacy_unmatched", "active": False, "requires_review": True,
                    "motivo": "no aparece en ninguna tabla del PDF de tarifas nuevo",
                })

    nuevos_codigos = sorted({f["product_code"] for f in filas
                             if f["tariff_code"] == "1"} - set(por_codigo))
    return {
        "codigos_catalogo_anterior": len(por_codigo),
        "precios_antiguos_con_correspondencia": resueltos,
        "precios_antiguos_huerfanos": huerfanos,
        "codigos_nuevos": nuevos_codigos,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--pdf")
    ap.add_argument("--xlsx", help="Fuente OFICIAL de precios.")
    ap.add_argument("--aprobar", type=int, metavar="N",
                    help="Aprueba y ACTIVA la versión N. Requiere --por.")
    ap.add_argument("--por", help="Quién aprueba. Queda en la auditoría.")
    ap.add_argument("--listar", action="store_true")
    a = ap.parse_args()

    TARIFAS.mkdir(parents=True, exist_ok=True)
    e = estado()

    if a.listar:
        print(f"Versión activa: {e['version_activa'] or '(ninguna)'}")
        for v in e["versiones"]:
            marca = "ACTIVA" if v["version"] == e["version_activa"] else \
                    ("aprobada" if v.get("approved") else "pendiente")
            print(f"  v{v['version']}  {v['generado'][:19]}  {v['registros']} reg  [{marca}]")
        return

    if a.aprobar is not None:
        if not a.por:
            sys.exit("Aprobar una tarifa exige --por: queda registrado quién lo hizo.")
        v = cargar_version(a.aprobar)
        if not v:
            sys.exit(f"No existe la versión {a.aprobar}.")
        if v["invariantes_fallidos"]:
            sys.exit(f"La versión {a.aprobar} no pasa sus invariantes: no se puede activar.")
        for f in v["filas"]:
            f["approved"] = True
            # Las especiales quedan importadas pero FUERA del flujo general:
            # una tarifa especial solo se usa por asociación explícita.
            f["active"] = f["tariff_code"] not in ESPECIALES
        v["approved"] = True
        v["approved_by"] = a.por
        v["approved_at"] = datetime.now(timezone.utc).isoformat()
        (TARIFAS / f"version-{a.aprobar}.json").write_text(
            json.dumps(v, ensure_ascii=False, indent=2), encoding="utf-8")

        e["version_activa"] = a.aprobar
        for x in e["versiones"]:
            if x["version"] == a.aprobar:
                x["approved"] = True
                x["approved_by"] = a.por
        e.setdefault("auditoria", []).append({
            "ts": datetime.now(timezone.utc).isoformat(), "accion": "activar_version",
            "version": a.aprobar, "por": a.por})
        guardar_estado(e)
        print(f"Versión {a.aprobar} aprobada y activada por {a.por}.")
        print("Las tarifas especiales (ALI, COO, OFC, S) quedan importadas pero INACTIVAS.")
        return

    if not a.pdf and not a.xlsx:
        sys.exit("Hace falta --xlsx, --pdf, --aprobar o --listar.")

    pdf = Path(a.xlsx or a.pdf).expanduser()
    sha = hashlib.sha256(pdf.read_bytes()).hexdigest()
    version = max([v["version"] for v in e["versiones"]], default=0) + 1

    filas, paginas = parsear_xlsx(pdf) if a.xlsx else parsear(pdf)
    fallos = comprobar_invariantes(filas)
    normalizadas = normalizar(filas, pdf, version, sha)

    activa = cargar_version(e["version_activa"]) if e["version_activa"] else None
    d = diff(normalizadas, activa["filas"] if activa else None)
    leg = legado(normalizadas)

    doc = {
        "version": version,
        "generado": datetime.now(timezone.utc).isoformat(),
        "source_file": pdf.name,
        "source_sha256": sha,
        "paginas": paginas,
        "registros": len(normalizadas),
        "approved": False,
        "active": False,
        "invariantes_fallidos": fallos,
        "resumen_por_tarifa": {k: sum(1 for f in normalizadas if f["tariff_code"] == k)
                               for k in TARIFAS_CONOCIDAS},
        "codigos_con_oferta": sorted({f["product_code"] for f in normalizadas
                                      if f["price_type"] == "offer"}),
        "diff_contra_activa": d,
        "legado": leg,
        "filas": normalizadas,
    }
    (TARIFAS / f"version-{version}.json").write_text(
        json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")

    e["versiones"].append({"version": version, "generado": doc["generado"],
                           "registros": doc["registros"], "approved": False,
                           "source_sha256": sha,
                           "invariantes_fallidos": len(fallos)})
    guardar_estado(e)

    origen = ("hoja «%s», %s filas" % (paginas["hoja"], paginas["filas_totales"])
              if paginas.get("origen") == "xlsx"
              else "%s páginas con contenido, %s en blanco"
                   % (paginas["paginas_con_contenido"], paginas["paginas_en_blanco"]))
    print(f"Versión {version} creada desde {pdf.name} · {len(normalizadas)} registros · {origen}")
    print()
    for k in TARIFAS_CONOCIDAS:
        print(f"  {k:5s} {doc['resumen_por_tarifa'][k]:4d}  {TARIFAS_CONOCIDAS[k][2]}")
    print()
    if fallos:
        print(f"INVARIANTES FALLIDOS ({len(fallos)}) — la versión NO se puede activar:")
        for f in fallos[:20]:
            print(f"  ✗ {f}")
    else:
        print("Invariantes: los 12 en verde.")
    print()
    print(f"Códigos con oferta: {len(doc['codigos_con_oferta'])} "
          f"({', '.join(doc['codigos_con_oferta'][:6])}…)")
    print(f"Códigos nuevos respecto al catálogo anterior: {len(leg.get('codigos_nuevos', []))}")
    print(f"Precios antiguos sin correspondencia: {len(leg.get('precios_antiguos_huerfanos', []))}")
    for h in leg.get("precios_antiguos_huerfanos", []):
        print(f"  {h['product_code']:8s} {h['price_display']:9s} -> legacy_unmatched, inactivo")
    print()
    if d["sin_version_anterior"]:
        print("Diff: no había versión anterior; todo es alta.")
    else:
        print(f"Diff contra la activa: {d['altas']} altas, {d['bajas']} bajas, "
              f"{len(d['cambios_de_precio'])} cambios de precio")
    print()
    print("La versión queda PENDIENTE. No cambia nada en producción hasta:")
    print(f"  python3 chacon-alcantara/import/extraer_tarifas.py --aprobar {version} --por 'Nombre'")


if __name__ == "__main__":
    main()
